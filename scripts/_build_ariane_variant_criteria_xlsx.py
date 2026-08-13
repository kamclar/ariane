from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path("outputs/019f8477-ada3-7743-b895-bf668851b5da/ariane_varianty_a_kriteria_2026-08-13.xlsx")


ROWS = [
    ["BRCA1", "c.181T>G", "c.181T>G", "p.(Cys61Gly)", "missense", "neuvedena", "PS3 Strong; PP3 Supporting; PP4 Very Strong", "Class 5", 13, "Doplněno z ARIANE", "Referenční báze T je správná."],
    ["BRCA1", "c.181C>G", "", "", "neplatný vstup", "žádná", "žádná", "Odmítnuto", None, "Odmítnuto správně", "Na c.181 je v referenčním transkriptu T, nikoliv C."],
    ["BRCA1", "c.5366C>T", "c.5366C>T", "p.(Ala1789Val)", "missense", "PP3", "PS3 Strong; PM2 Supporting", "Class 3", 5, "Neshoda", "Podklad očekává PP3, ale aktuální ARIANE jej neaplikuje."],
    ["BRCA1", "c.5542C>T", "c.5542C>T", "p.(Gln1848Ter)", "nonsense", "PM5 Strong", "PVS1 Very Strong; PS3 Strong; PM5-PTC Strong; PM2 Supporting", "Class 5", 17, "Shoda po opravě", "PM5-PTC Strong je nyní aplikováno."],
    ["BRCA1", "c.3247A>C", "c.3247A>C", "p.(Met1083Leu)", "missense", "BP1 Strong", "PP4 Moderate", "Class 3", 2, "Neshoda", "Podklad očekává BP1 Strong, ARIANE aplikuje PP4 Moderate."],
    ["BRCA1", "c.3247A>G", "c.3247A>G", "p.(Met1083Val)", "missense", "BP1 Strong, ale v textu spojeno s p.(Met1083Leu)", "BP1 Strong", "Class 2", -4, "Oprava zápisu", "c.3247A>G vede k Met1083Val; Met1083Leu patří k c.3247A>C."],
    ["BRCA1", "c.5217T>A", "c.5217T>A", "p.(Asp1739Glu)", "missense", "PS3 Strong; PM2 Supporting; PP3 Supporting", "PS3 Strong; PM2 Supporting; PP3 Supporting", "Class 4", 6, "Shoda", "Kritéria se shodují s podkladem."],
    ["BRCA2", "c.3703C>T", "c.3703C>T", "p.(Gln1235Ter)", "nonsense", "neuvedena", "PVS1 Very Strong; PM5-PTC Strong; PM2 Supporting", "Class 5", 13, "Doplněno z ARIANE", "V podkladu šlo o test nerozlišování velikosti písmen."],
    ["BRCA1", "c.3668_3671dup", "c.3668_3671dup", "p.(Cys1225SerfsTer10)", "duplikace, frameshift", "neuvedena", "PVS1 Very Strong; PM5-PTC Strong; PP4 Strong", "Class 5", 16, "Doplněno z ARIANE", "ARIANE doplnila plnou třípísmennou p. notaci."],
    ["BRCA1", "c.2102delA / c.2102del", "c.2102del", "p.(Lys701SerfsTer2)", "delece, frameshift", "neuvedena", "PVS1 Very Strong; PM5-PTC Strong", "Class 5", 12, "Nyní funkční", "Oba vstupy se normalizují na c.2102del."],
    ["BRCA1", "c.3141_3161del", "c.3141_3161del", "p.(Thr1051_Ser1057del)", "in-frame delece", "neuvedena", "žádná", "Class 3", 0, "Doplněno z ARIANE", "V podkladu bylo uvedeno pouze to, že vstup funguje."],
    ["BRCA1", "c.3877_3878delinsTT", "c.3877_3878delinsTT", "p.(Ala1293Phe)", "delins, missense", "neuvedena", "žádná", "Class 3", 0, "Doplněno z ARIANE", "Příklad podporovaného zápisu."],
    ["BRCA1", "c.5569C>T", "c.5569C>T", "p.(Gln1857Ter)", "nonsense", "neuvedena", "PM2 Supporting", "Class 3", 1, "Doplněno z ARIANE", "Příklad podporovaného zápisu."],
    ["BRCA1", "c.5193G>A", "c.5193G>A", "p.(Glu1731=)", "synonymní", "neuvedena", "PM2 Supporting; BS3 Strong; BP4 Supporting; BP7 Supporting", "Class 2", -5, "Doplněno z ARIANE", "ARIANE kanonizovala starší synonymní zápis na p.(Glu1731=)."],
    ["BRCA1", "c.5478_5479dup", "c.5478_5479dup", "p.(Met1827ArgfsTer8)", "duplikace, frameshift", "neuvedena", "PVS1 Very Strong; PM5-PTC Strong", "Class 5", 12, "Doplněno z ARIANE", "ARIANE přijímá zkrácené p. aliasy a vrací plnou notaci."],
    ["BRCA1", "c.5213_5215del", "c.5213_5215del", "p.(Gly1738del)", "in-frame delece", "neuvedena", "PP4 Supporting", "Class 3", 1, "Doplněno z ARIANE", "Příklad podporovaného zápisu."],
    ["BRCA1", "c.5467+2T>G", "c.5467+2T>G", "p.?", "splice substituce", "neuvedena", "PVS1 Very Strong; PM2 Supporting", "Class 4", 9, "Doplněno z ARIANE", "Proteinový následek nelze bez RNA bezpečně určit."],
    ["BRCA1", "c.5556_5560del", "c.5556_5560del", "p.(Tyr1853AspfsTer25)", "delece, frameshift", "PVS1 Very Strong; PM5 Strong", "PVS1 Very Strong; PM5-PTC Strong", "Class 5", 12, "Shoda po opravě", "Oprava podle Appendix D je aktivní."],
    ["BRCA1", "c.5533_5534insG", "c.5533_5534insG", "p.(Tyr1845Ter)", "inzerce s PTC", "PVS1 Very Strong; PM5 Strong", "PVS1 Very Strong; PM5-PTC Strong", "Class 5", 12, "Shoda po opravě", "PM2 se pro tento indel již nepoužívá."],
    ["BRCA2", "c.9891_9894dup", "c.9891_9894dup", "p.(Gln3299IlefsTer29)", "duplikace, frameshift", "PVS1 Very Strong; PM5 Strong; možná BS1 Supporting", "PVS1 Very Strong; PM5-PTC Strong; BP5 Supporting", "Class 5", 11, "Částečná shoda", "PVS1 a PM5 se shodují; ARIANE přidává BP5 Supporting, nikoliv BS1."],
]


HEADERS = [
    "Gen",
    "Vstupní c. notace",
    "Kanonická c. notace",
    "p. následek ARIANE",
    "Typ varianty",
    "Kritéria v podkladu",
    "Kritéria ARIANE",
    "Klasifikace ARIANE",
    "Body",
    "Stav porovnání",
    "Poznámka",
]


def make_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Varianty"
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "2F75B5"
    pale_blue = "DCE6F1"
    pale_green = "E2F0D9"
    pale_yellow = "FFF2CC"
    pale_red = "FCE4D6"
    pale_gray = "E7E6E6"
    dark = "1F2937"
    white = "FFFFFF"
    border_color = "CCD6E0"
    thin = Side(style="thin", color=border_color)

    ws.merge_cells("A1:K1")
    ws["A1"] = "ARIANE: varianty, klasifikace a použitá kritéria"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=white)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    ws["A2"] = "Porovnání údajů z dodaného textu s aktuálním lokálním výstupem ARIANE dne 13. 8. 2026"
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="52606D")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    cards = [
        ("A4:B4", "A5:B5", "Počet řádků", "=COUNTA(A8:A27)", pale_blue),
        ("C4:D4", "C5:D5", "Class 5", '=COUNTIF(H8:H27,"Class 5")', pale_red),
        ("E4:F4", "E5:F5", "Neshody", '=COUNTIF(J8:J27,"Neshoda*")', pale_yellow),
        ("G4:H4", "G5:H5", "Odmítnuté vstupy", '=COUNTIF(H8:H27,"Odmítnuto")', pale_gray),
        ("I4:K4", "I5:K5", "Datum kontroly", "2026-08-13", pale_green),
    ]
    for label_range, value_range, label, value, fill in cards:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = value
        label_cell.fill = PatternFill("solid", fgColor=blue)
        label_cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = PatternFill("solid", fgColor=fill)
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color=dark)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws[label_range]:
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws[value_range]:
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28

    for col, value in enumerate(HEADERS, start=1):
        cell = ws.cell(row=7, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[7].height = 34

    for row_idx, row in enumerate(ROWS, start=8):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Aptos", size=10, color=dark)
            cell.alignment = Alignment(
                horizontal="right" if col_idx == 9 else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row_idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F7F9FC")
        ws.cell(row=row_idx, column=1).font = Font(name="Aptos", size=10, bold=True, color=dark)
        ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row_idx, column=9).number_format = "0;[Red]-0;0"
        ws.row_dimensions[row_idx].height = 54

    table = Table(displayName="ArianeVariantCriteria", ref="A7:K27")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = "A7:K27"
    widths = {"A": 10, "B": 24, "C": 22, "D": 27, "E": 23, "F": 38, "G": 52, "H": 20, "I": 9, "J": 24, "K": 50}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.conditional_formatting.add("J8:J27", FormulaRule(formula=['LEFT(J8,7)="Neshoda"'], fill=PatternFill("solid", fgColor="F4CCCC")))
    ws.conditional_formatting.add("J8:J27", FormulaRule(formula=['OR(LEFT(J8,5)="Shoda",LEFT(J8,4)="Nyní")'], fill=PatternFill("solid", fgColor="D9EAD3")))
    ws.conditional_formatting.add("J8:J27", FormulaRule(formula=['LEFT(J8,9)="Odmítnuto"'], fill=PatternFill("solid", fgColor="FCE5CD")))
    ws.conditional_formatting.add("J8:J27", FormulaRule(formula=['OR(LEFT(J8,9)="Částečná",LEFT(J8,6)="Oprava")'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add("H8:H27", FormulaRule(formula=['H8="Class 5"'], fill=PatternFill("solid", fgColor="F4CCCC")))
    ws.conditional_formatting.add("H8:H27", FormulaRule(formula=['H8="Class 4"'], fill=PatternFill("solid", fgColor="FCE5CD")))
    ws.conditional_formatting.add("H8:H27", FormulaRule(formula=['H8="Class 3"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add("H8:H27", FormulaRule(formula=['H8="Class 2"'], fill=PatternFill("solid", fgColor="D9EAD3")))

    notes = wb.create_sheet("Poznámky")
    notes.sheet_view.showGridLines = False
    notes.merge_cells("A1:F1")
    notes["A1"] = "Poznámky k porovnání"
    notes["A1"].fill = PatternFill("solid", fgColor=navy)
    notes["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=white)
    notes["A1"].alignment = Alignment(horizontal="center")
    notes.row_dimensions[1].height = 32
    note_rows = [
        ["Rozsah", "Tabulka obsahuje unikátní varianty a testovací vstupy vybrané z dodaného textu. Kritéria ARIANE pocházejí z aktuálního lokálního běhu."],
        ["Aplikovaná kritéria", "Uvedena jsou pouze kritéria, která ARIANE skutečně aplikovala a započítala. Kandidáti pro manuální revizi nejsou ve sloupci Kritéria ARIANE."],
        ["BRCA1 c.5366C>T", "Podklad očekává PP3, ale aktuální ARIANE vrací PS3 Strong a PM2 Supporting bez PP3. Tato neshoda zůstává k prověření."],
        ["BRCA1 c.3247A>C", "Tato varianta vede k p.(Met1083Leu). Podklad očekává BP1 Strong, zatímco aktuální ARIANE vrací PP4 Moderate."],
        ["BRCA1 c.3247A>G", "Jde o jinou variantu s následkem p.(Met1083Val). ARIANE u ní vrací BP1 Strong."],
        ["Referenční kontrola", "BRCA1 c.181C>G je správně odmítnuta, protože referenční báze na c.181 je T."],
        ["Časová platnost", "Výsledky odpovídají lokálnímu stavu aplikace a dostupným datovým zdrojům dne 13. 8. 2026."],
    ]
    notes["A3"] = "Téma"
    notes["B3"] = "Vysvětlení"
    for cell in notes[3]:
        if cell.column <= 2:
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(name="Aptos", bold=True, color=white)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for idx, row in enumerate(note_rows, start=4):
        notes.cell(idx, 1, row[0])
        notes.cell(idx, 2, row[1])
        notes.cell(idx, 1).font = Font(name="Aptos", bold=True, color=dark)
        for col in (1, 2):
            notes.cell(idx, col).alignment = Alignment(vertical="top", wrap_text=True)
            notes.cell(idx, col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if idx % 2 == 1:
                notes.cell(idx, col).fill = PatternFill("solid", fgColor="F7F9FC")
        notes.row_dimensions[idx].height = 48
    notes.column_dimensions["A"].width = 26
    notes.column_dimensions["B"].width = 105
    notes.freeze_panes = "A4"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    assert check.sheetnames == ["Varianty", "Poznámky"]
    assert check["Varianty"].max_row == 27
    assert check["Varianty"]["G10"].value == "PS3 Strong; PM2 Supporting"
    assert check["Varianty"]["H13"].value == "Class 2"
    assert check["Varianty"]["A27"].value == "BRCA2"
    print(OUTPUT.resolve())


if __name__ == "__main__":
    make_workbook()
