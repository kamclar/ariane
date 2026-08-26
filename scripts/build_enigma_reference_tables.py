"""Build a deterministic browser snapshot of all ENIGMA v1.2 tables.

The source bundle is checksum-validated separately by
``verify_enigma_source_bundle.py``. This builder extracts three distinct table
series without retyping their scientific contents:

* Specifications Tables 1 to 9
* Appendix Tables 1 to 17
* Supplementary Tables ST1 to ST16

The resulting JSON is a read-only presentation snapshot. Runtime classifier
lookups continue to use their dedicated, validated datasets such as
``enigma_table4.json`` and ``enigma_table9.json``.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime
import hashlib
import json
import math
from pathlib import Path
import re
import xml.etree.ElementTree as ET
import zipfile

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
BUNDLE = ROOT / "docs" / "enigma" / "v1.2"
SOURCE = BUNDLE / "source"
MANIFEST = BUNDLE / "manifest.json"

WORD_NAMESPACE = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}

GROUPS = [
    {
        "id": "specification",
        "title": "Specification Tables",
        "description": "Tables 1 to 9 in the ENIGMA BRCA1/2 VCEP v1.2 specification.",
    },
    {
        "id": "appendix",
        "title": "Appendix Tables",
        "description": "Tables 1 to 17 in the ENIGMA BRCA1/2 VCEP v1.2 Appendix.",
    },
    {
        "id": "supplementary",
        "title": "Supplementary Tables",
        "description": "Supplementary Tables ST1 to ST16 supplied in the official workbook.",
    },
]

USAGE = {
    "specification-table-1": ("rule_definition", "Definitions and gene-specific application of ACMG/AMP criteria."),
    "specification-table-2": ("rule_definition", "Evidence-strength odds used by ENIGMA criteria."),
    "specification-table-3": ("runtime_rule", "Default combinations used by the classifier."),
    "specification-table-4": ("runtime_lookup", "PVS1 and PM5 PTC decisions queried by the classifier."),
    "specification-table-5": ("expert_review", "PS1 splicing weights used during expert review."),
    "specification-table-6": ("expert_review", "PM3 point assignment used during expert review."),
    "specification-table-7": ("rule_definition", "PP4 and BP5 examples supporting clinical LR interpretation."),
    "specification-table-8": ("expert_review", "BS2 point assignment used during expert review."),
    "specification-table-9": ("runtime_lookup", "PS3, BS3 and curated splice information queried by the classifier."),
    "appendix-table-3": ("runtime_rule", "BRCA1 domains and clinically important residues."),
    "appendix-table-4": ("runtime_rule", "BRCA2 domains and clinically important residues."),
    "appendix-table-9": ("runtime_rule", "RNA evidence weighting implemented by the RNA pathway."),
    "appendix-table-11": (
        "runtime_lookup",
        "Variant-specific moderate-risk and reduced-penetrance annotations queried by ARIANE.",
    ),
    "appendix-table-14": ("rule_definition", "Missense prior probabilities by domain and prediction."),
    "appendix-table-15": ("rule_definition", "Calibration of missense bioinformatic likelihood ratios."),
    "appendix-table-16": ("runtime_rule", "Bioinformatic and domain code recommendations."),
    "appendix-table-17": ("expert_review", "PS1 splicing weights used during expert review."),
    "supplementary-table-2": ("runtime_lookup", "Curated mRNA assay results queried by the RNA and splice-review pathways."),
    "supplementary-table-7": ("candidate_registry", "Trusted P/LP reference candidates used by the protein PS1 review workflow."),
}

DEFAULT_USAGE = {
    "specification": ("rule_definition", "Normative ENIGMA rule or supporting example."),
    "appendix": ("rule_support", "Mechanism-specific justification, calibration or supporting evidence."),
    "supplementary": ("calibration_reference", "Official calibration or reference dataset, not directly queried unless stated otherwise."),
}


def _manifest_files() -> dict[str, dict]:
    payload = json.loads(MANIFEST.read_text(encoding="utf-8"))
    return {Path(item["path"]).name: item for item in payload["files"]}


def _verify_sources(files: dict[str, dict]) -> None:
    errors: list[str] = []
    for filename, item in files.items():
        path = SOURCE / filename
        if not path.is_file():
            errors.append(f"missing source: {filename}")
            continue
        if path.stat().st_size != item["size_bytes"]:
            errors.append(f"size mismatch: {filename}")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest != item["sha256"]:
            errors.append(f"checksum mismatch: {filename}")
    if errors:
        raise RuntimeError("ENIGMA source validation failed: " + "; ".join(errors))


def _word_text(element: ET.Element) -> str:
    return "".join(
        node.text or "" for node in element.findall(".//w:t", WORD_NAMESPACE)
    ).strip()


def _trim_grid(rows: list[list[object]]) -> list[list[object]]:
    while rows and not any(_cell_present(value) for value in rows[-1]):
        rows.pop()
    last_column = 0
    for row in rows:
        for index, value in enumerate(row, 1):
            if _cell_present(value):
                last_column = max(last_column, index)
    if last_column == 0:
        return []
    return [row[:last_column] + [None] * max(0, last_column - len(row)) for row in rows]


def _cell_present(value: object) -> bool:
    if isinstance(value, dict):
        return bool(value.get("formula")) or value.get("value") not in {None, ""}
    return value not in {None, ""}


def _clean_cell(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _extract_docx_tables(path: Path, expected) -> dict[int, dict]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
    body = root.find("w:body", WORD_NAMESPACE)
    if body is None:
        raise RuntimeError(f"DOCX body missing: {path.name}")

    expected_numbers = set(expected)
    found: dict[int, dict] = {}
    pending: tuple[int, str] | None = None
    for element in body:
        tag = element.tag.rsplit("}", 1)[-1]
        if tag == "p":
            text = _word_text(element)
            match = re.match(r"^Table\s+(\d+)\s*:\s*(.+)$", text, re.IGNORECASE)
            if match:
                number = int(match.group(1))
                pending = (number, text) if number in expected_numbers else None
        elif tag == "tbl" and pending:
            number, caption = pending
            rows = [
                [_clean_cell(_word_text(cell)) for cell in row.findall("./w:tc", WORD_NAMESPACE)]
                for row in element.findall("./w:tr", WORD_NAMESPACE)
            ]
            rows = _trim_grid(rows)
            if rows:
                # Later occurrences replace table-of-contents artefacts.
                found[number] = {"caption": caption, "rows": rows}
            pending = None

    missing = expected_numbers - set(found)
    if missing:
        raise RuntimeError(f"Missing tables in {path.name}: {sorted(missing)}")
    return found


def _extract_xlsx_sheet(
    values_sheet,
    formula_sheet,
) -> list[list[object]]:
    rows: list[list[object]] = []
    value_rows = values_sheet.iter_rows(values_only=True)
    formula_rows = formula_sheet.iter_rows(values_only=True)
    for values, formulas in zip(value_rows, formula_rows):
        row: list[object] = []
        for value, raw in zip(values, formulas):
            cleaned = _clean_cell(value)
            if isinstance(raw, str) and raw.startswith("="):
                row.append({"value": cleaned, "formula": raw})
            else:
                row.append(_clean_cell(raw if cleaned is None else cleaned))
        rows.append(row)
    return _trim_grid(rows)


def _extract_xlsx(path: Path) -> list[dict]:
    values_book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    formula_book = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sections: list[dict] = []
    for index, values_sheet in enumerate(values_book.worksheets, 1):
        formula_sheet = formula_book[values_sheet.title]
        rows = _extract_xlsx_sheet(values_sheet, formula_sheet)
        sections.append({
            "id": f"sheet-{index}",
            "title": values_sheet.title,
            "row_count": len(rows),
            "column_count": max((len(row) for row in rows), default=0),
            "rows": rows,
        })
    values_book.close()
    formula_book.close()
    return sections


def _table_record(
    *,
    table_id: str,
    group: str,
    number: int,
    title: str,
    source_id: str,
    source_file: dict,
    sections: list[dict],
) -> dict:
    usage_id, usage_text = USAGE.get(table_id, DEFAULT_USAGE[group])
    return {
        "id": table_id,
        "group": group,
        "number": number,
        "title": title,
        "source_id": source_id,
        "source_sha256": source_file["sha256"],
        "usage": usage_id,
        "usage_text": usage_text,
        "sections": sections,
    }


def _inline_section(rows: list[list[object]]) -> list[dict]:
    return [{
        "id": "main",
        "title": "Main table",
        "row_count": len(rows),
        "column_count": max((len(row) for row in rows), default=0),
        "rows": rows,
    }]


def build_snapshot() -> dict:
    files = _manifest_files()
    _verify_sources(files)
    specification_file = files["Specifications_V1.2.docx"]
    appendix_file = files["Appendix_V1.2.docx"]
    table4_file = files["Specifications_Table4_V1.2.xlsx"]
    table9_file = files["Specifications_Table9_V1.2.xlsx"]
    supplementary_file = files["Supplementary_Tables_V1.2.xlsx"]

    specification_inline = _extract_docx_tables(
        SOURCE / "Specifications_V1.2.docx",
        (1, 2, 3, 5, 6, 7, 8),
    )

    appendix_inline = _extract_docx_tables(
        SOURCE / "Appendix_V1.2.docx",
        range(1, 18),
    )

    table4_sections = _extract_xlsx(SOURCE / "Specifications_Table4_V1.2.xlsx")
    table9_sections = _extract_xlsx(SOURCE / "Specifications_Table9_V1.2.xlsx")
    supplementary_sections = _extract_xlsx(SOURCE / "Supplementary_Tables_V1.2.xlsx")
    if len(supplementary_sections) != 16:
        raise RuntimeError("Expected 16 ENIGMA supplementary table worksheets")

    tables: list[dict] = []
    for number in range(1, 10):
        table_id = f"specification-table-{number}"
        if number == 4:
            title = "Table 4: PVS1 decision-tree and PM5 PTC codes"
            sections = table4_sections
            source_id = "enigma-v1.2-table4"
            source_file = table4_file
        elif number == 9:
            title = "Table 9: reviewed functional assay results for PS3 and BS3"
            sections = table9_sections
            source_id = "enigma-v1.2-table9"
            source_file = table9_file
        else:
            item = specification_inline[number]
            title = item["caption"]
            sections = _inline_section(item["rows"])
            source_id = "enigma-v1.2-specifications"
            source_file = specification_file
        tables.append(_table_record(
            table_id=table_id,
            group="specification",
            number=number,
            title=title,
            source_id=source_id,
            source_file=source_file,
            sections=sections,
        ))

    for number in range(1, 18):
        item = appendix_inline[number]
        tables.append(_table_record(
            table_id=f"appendix-table-{number}",
            group="appendix",
            number=number,
            title=item["caption"],
            source_id="enigma-v1.2-appendix",
            source_file=appendix_file,
            sections=_inline_section(item["rows"]),
        ))

    for number, section in enumerate(supplementary_sections, 1):
        first_text = next(
            (
                value for row in section["rows"] for value in row
                if isinstance(value, str) and value.strip()
            ),
            f"Supplementary Table {number}",
        )
        if not first_text.lower().startswith("supplementary table"):
            sheet_label = re.sub(
                rf"^ST{number}\s*",
                "",
                section["title"],
                flags=re.IGNORECASE,
            ).strip()
            first_text = f"Supplementary Table {number}: {sheet_label}"
        section["id"] = "main"
        tables.append(_table_record(
            table_id=f"supplementary-table-{number}",
            group="supplementary",
            number=number,
            title=first_text,
            source_id="enigma-v1.2-supplementary-tables",
            source_file=supplementary_file,
            sections=[section],
        ))

    expected_ids = {
        *(f"specification-table-{number}" for number in range(1, 10)),
        *(f"appendix-table-{number}" for number in range(1, 18)),
        *(f"supplementary-table-{number}" for number in range(1, 17)),
    }
    actual_ids = {table["id"] for table in tables}
    if actual_ids != expected_ids or len(tables) != 42:
        raise RuntimeError("ENIGMA table inventory is incomplete")

    return {
        "schema_version": 1,
        "criteria_specification_id": "GN092",
        "criteria_specification_version": "1.2.0",
        "release_date": "2025-01-09",
        "view_description": (
            "Structured presentation of checksum-validated ENIGMA source tables. "
            "Merged cells and visual formatting may be simplified; the linked "
            "official documents remain authoritative."
        ),
        "groups": GROUPS,
        "table_count": len(tables),
        "tables": tables,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT / "backend" / "data" / "enigma_reference_tables.json",
    )
    args = parser.parse_args()
    snapshot = build_snapshot()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {snapshot['table_count']} ENIGMA tables to {args.output}")


if __name__ == "__main__":
    main()
