"""Build a lossless, executable ENIGMA BRCA1/2 VCEP Table 4 snapshot."""
from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

import openpyxl


SOURCE_URL = (
    "https://cspec.genome.network/cspec/File/id/"
    "ca5cf57b-94df-4ad6-a001-c62ceccb3845/data"
)


def _clean(value):
    if isinstance(value, str):
        return value.strip()
    return value


def _position(value) -> int | None:
    if value is None:
        return None
    match = re.fullmatch(r"c\.(-?\d+)\*?", str(value).strip())
    return int(match.group(1)) if match else None


def _notes(*values) -> str:
    return " | ".join(str(value).strip() for value in values if value not in (None, ""))


def _alleles(value) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _add_splice_rules(target, gene, exon, pattern, alleles, code, notes):
    if not pattern or not code:
        return
    pattern = str(pattern).strip()
    if not pattern.startswith("c."):
        raise RuntimeError(f"Unexpected Table 4 splice pattern: {pattern!r}")
    base = pattern[:-1] if pattern.endswith(">") else pattern
    for allele in _alleles(alleles):
        key = base + allele
        if key in target[gene]:
            previous = target[gene][key]
            new_code = str(code).strip()
            # Table 4 deliberately lists a generic prediction and then a
            # variant-specific RNA result for a few alleles.  The latter is
            # the executable rule while both rows remain in source_rows.
            if "(RNA)" not in new_code and previous["pvs1_code"] != new_code:
                raise RuntimeError(f"Conflicting Table 4 splice rule: {gene}:{key}")
        target[gene][key] = {
            "exon": exon,
            "pvs1_code": str(code).strip(),
            "notes": _notes(notes),
        }


def build_snapshot(source: Path) -> dict:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    readme_sheet = workbook["Specifications - Table 4 Readme"]
    sheet = workbook["Table 4 - Annotated Exons"]
    rows = [[_clean(value) for value in row] for row in sheet.iter_rows(values_only=True)]

    exon_ranges = {"BRCA1": {}, "BRCA2": {}}
    ptc_rules = {"BRCA1": {}, "BRCA2": {}}
    splice_rules = {"BRCA1": {}, "BRCA2": {}}
    deletion_rules = {"BRCA1": {}, "BRCA2": {}}
    duplication_rules = {"BRCA1": {}, "BRCA2": {}}
    critical_boundaries = {}

    for row_number, row in enumerate(rows[1:], 2):
        gene, exon = row[0], row[1]
        if gene not in {"BRCA1", "BRCA2"} or not exon:
            continue

        start = next(
            (position for index in (9, 8, 7) if (position := _position(row[index])) is not None),
            None,
        )
        end = next(
            (position for index in (11, 12, 13) if (position := _position(row[index])) is not None),
            None,
        )
        if start is not None and end is not None:
            exon_ranges[gene][exon] = [start, end]

        marker = str(row[10]).strip() if row[10] is not None else ""
        code = row[9]
        if marker == "PTC":
            ptc_rules[gene][exon] = {
                "pvs1_code": code,
                "pm5_code": row[11],
                "condition": None,
            }
        elif marker.startswith("PTC<"):
            boundary_match = re.search(r"p\.[A-Z][a-z]{0,2}(\d+)", marker)
            if not boundary_match:
                raise RuntimeError(f"Cannot parse PTC boundary at row {row_number}: {marker}")
            first_na_aa = int(boundary_match.group(1))
            critical_boundaries[gene] = {
                "exon": exon,
                "boundary_aa": first_na_aa - 1,
                "rule": marker,
            }
        elif marker.startswith("PTC>"):
            ptc_rules[gene][exon] = {
                "pvs1_code": code,
                "pm5_code": row[11],
                "condition": marker,
            }
            if gene in critical_boundaries:
                critical_boundaries[gene]["rule"] += f"; {marker}"
        elif marker.startswith("DEL "):
            deletion_rules[gene][exon] = {
                "pvs1_code": code,
                "notes": _notes(row[2], row[18]),
            }
        elif marker.startswith("DUP ") or marker == "DUP 11":
            arrangement = str(row[11]).strip()
            duplication_rules[gene].setdefault(exon, {})[arrangement] = {
                "pvs1_code": code,
                "notes": _notes(row[2], row[18]),
            }

        _add_splice_rules(
            splice_rules, gene, exon,
            row[4], row[5], row[3], row[2],
        )
        _add_splice_rules(
            splice_rules, gene, exon,
            row[15], row[16], row[17], row[18],
        )

    for gene in ("BRCA1", "BRCA2"):
        expected = set(exon_ranges[gene])
        for name, section in (
            ("deletion", deletion_rules),
            ("duplication", duplication_rules),
        ):
            missing = expected - set(section[gene])
            if missing:
                raise RuntimeError(f"Missing {name} rules for {gene}: {sorted(missing)}")

    readme = [_clean(row[0]) for row in readme_sheet.iter_rows(values_only=True)]
    return {
        "schema_version": 2,
        "version": "1.2.0",
        "released": "2025-01-09",
        "source": "ClinGen ENIGMA BRCA1/2 VCEP Specifications Table 4",
        "source_url": SOURCE_URL,
        "generated": date.today().isoformat(),
        "readme": readme,
        "source_columns": 20,
        "source_rows": rows,
        "exon_ranges": exon_ranges,
        "ptc_rules": ptc_rules,
        "splice_rules": splice_rules,
        "deletion_rules": deletion_rules,
        "duplication_rules": duplication_rules,
        "critical_boundaries": critical_boundaries,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument(
        "--output", type=Path, default=Path("backend/data/enigma_table4.json")
    )
    args = parser.parse_args()
    snapshot = build_snapshot(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"Wrote lossless Table 4 snapshot with {len(snapshot['source_rows'])} source rows "
        f"to {args.output}"
    )


if __name__ == "__main__":
    main()
