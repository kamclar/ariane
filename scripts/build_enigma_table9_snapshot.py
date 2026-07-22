"""Build the lossless ENIGMA BRCA1/2 VCEP Table 9 JSON snapshot.

The input is the official ``Specifications_Table9_V1.2.xlsx`` workbook from
ClinGen CSpec.  Unlike the original application snapshot, this preserves all
reviewed rows (including rows where PS3/BS3 is not met) and every source
column relevant to protein-function and splicing interpretation.
"""
from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

import openpyxl


EXPECTED_COLUMNS = [
    "Gene",
    "HGVS Nucleotide",
    "HGVS Protein",
    "Assigned Code",
    "Code Weight",
    "Standardised Text",
    "Splice Result Published",
    "Splicing Prediction",
    "Predicted or Observed Splicing",
    "# Pubs",
    "Result1",
    "Result2",
    "Result3",
    "Result4",
]

FIELD_NAMES = [
    "gene",
    "c_notation",
    "p_notation",
    "code",
    "strength",
    "text",
    "splice_result_published",
    "spliceai_prediction",
    "predicted_or_observed_splicing",
    "publication_count",
    "result_1",
    "result_2",
    "result_3",
    "result_4",
]


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def build_snapshot(source: Path) -> dict:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Table9_BRCA12VCEP_specs"]

    header_row = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if list(row[: len(EXPECTED_COLUMNS)]) == EXPECTED_COLUMNS:
            header_row = row_number
            break
    if header_row is None:
        raise RuntimeError("Official Table 9 header was not found or has changed")

    variants = {}
    for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        values = [_clean(value) for value in row[: len(FIELD_NAMES)]]
        if not any(value is not None for value in values):
            continue
        entry = dict(zip(FIELD_NAMES, values))
        gene = entry["gene"]
        c_notation = entry["c_notation"]
        if gene not in {"BRCA1", "BRCA2"} or not str(c_notation).startswith("c."):
            raise RuntimeError(f"Invalid Table 9 row: {values!r}")
        key = f"{gene}:{c_notation}"
        if key in variants:
            raise RuntimeError(f"Duplicate Table 9 variant: {key}")
        variants[key] = entry

    return {
        "schema_version": 2,
        "version": "1.2.0",
        "released": "2025-01-09",
        "source": "ClinGen ENIGMA BRCA1/2 VCEP Specifications Table 9",
        "source_url": (
            "https://cspec.genome.network/cspec/File/id/"
            "0a35d6a8-5050-44b6-8a9d-babe8cdc06b2/data"
        ),
        "generated": date.today().isoformat(),
        "row_count": len(variants),
        "variants": variants,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("backend/data/enigma_table9.json"),
    )
    args = parser.parse_args()
    snapshot = build_snapshot(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {snapshot['row_count']} Table 9 rows to {args.output}")


if __name__ == "__main__":
    main()
