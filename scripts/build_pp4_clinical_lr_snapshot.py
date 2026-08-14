"""Build the local BRCA1/2 PP4/BP5 combined clinical-LR snapshot.

The build combines two versioned public ENIGMA resources:

* multifactorial clinical LRs from the UCSC ENIGMA ``BRCAmfa`` track;
* case-control LRs from Zanti et al. 2025 Supplementary Data 5.

Every component remains visible in provenance.  A variant receives one final
PP4 or BP5 code from the product of the admitted components; component-level
codes are never scored separately.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
DEFAULT_SOURCE = ROOT / "data/sources/enigma/BRCAmfa.hg38.v1.1.bed"
DEFAULT_CASE_CONTROL_SOURCE = (
    ROOT / "data/sources/enigma/Zanti_2025_NatCommun_Supplementary_Data.xlsx"
)
DEFAULT_SOURCE_MANIFEST = ROOT / "data/sources/enigma/clinical_lr_sources.manifest.json"
DEFAULT_OUTPUT = ROOT / "data/precomputed/brca_pp4_clinical_lr_snapshot.index.json"
DEFAULT_METADATA = ROOT / "data/precomputed/brca_pp4_clinical_lr_snapshot.metadata.json"
INDEL_INDEX = ROOT / "data/precomputed/brca_normalized_indel_snapshot.index.json"
INDEL_METADATA = ROOT / "data/precomputed/brca_normalized_indel_snapshot.metadata.json"
TRANSCRIPTS = {"NM_007294.4": "BRCA1", "NM_000059.4": "BRCA2"}
TRACK_URL = "https://hgdownload.soe.ucsc.edu/hubs/enigma/hg38/BRCAmfa.bb"
TRACK_DESCRIPTION_URL = "https://hgdownload.soe.ucsc.edu/hubs/enigma/enigma.html"
ZANTI_SOURCE_URL = (
    "https://media.springernature.com/original/springer-static/esm/"
    "art%3A10.1038%2Fs41467-025-59979-6/MediaObjects/"
    "41467_2025_59979_MOESM4_ESM.xlsx"
)
ZANTI_SHEET = "Supplementary Data 5"

# Field order is Family history, Co-occurrence, Segregation, Pathology,
# Case-control, as documented by the UCSC item schema/description.
SOURCES = {
    "caputoLRs": {
        "source_id": "caputo_2021_cosegregation",
        "citation": "Caputo et al. 2021",
        "pmid": "34597585",
        "doi": "10.1016/j.ajhg.2021.09.003",
        "data_types": ["family_history", "cooccurrence", "segregation", "pathology"],
        "independence_group": "caputo_2021_clinical_cohort",
    },
    "parsonsLRs": {
        "source_id": "parsons_2019_multifactorial",
        "citation": "Parsons et al. 2019",
        "pmid": "31131967",
        "doi": "10.1002/humu.23818",
        "data_types": ["family_history", "cooccurrence", "segregation", "pathology", "case_control"],
        "independence_group": "parsons_2019_multifactorial_dataset",
    },
    "liLRs": {
        "source_id": "li_2020_personal_family_history",
        "citation": "Li et al. 2020",
        "pmid": "31853058",
        "doi": "10.1038/s41436-019-0729-3",
        "data_types": ["personal_and_family_history"],
        "independence_group": "li_2020_ambry_testing_cohort",
    },
    "eastonLRs": {
        "source_id": "easton_2007_multifactorial",
        "citation": "Easton et al. 2007",
        "pmid": "17924331",
        "doi": "10.1086/521032",
        "data_types": ["family_history", "cooccurrence", "segregation", "pathology", "case_control"],
        "independence_group": "easton_2007_multifactorial_dataset",
    },
}

ZANTI_SOURCE = {
    "source_id": "zanti_2025_case_control",
    "citation": "Zanti et al. 2025",
    "pmid": "40413188",
    "doi": "10.1038/s41467-025-59979-6",
    "data_types": ["case_control"],
    "independence_group": "zanti_2025_bridges_carriers_ukb",
    "cohorts": ["BRIDGES/BCAC", "CARRIERS", "UK Biobank"],
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_source_manifest(
    manifest_path: Path,
    multifactorial_source: Path,
    case_control_source: Path,
) -> dict:
    """Load the source contract and verify both immutable build inputs."""
    if not manifest_path.is_file():
        raise RuntimeError(f"Clinical LR source manifest is missing: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("status") != "validated_source_manifest":
        raise RuntimeError("Clinical LR source manifest is not validated")
    configured = manifest.get("datasets")
    if not isinstance(configured, dict):
        raise RuntimeError("Clinical LR source manifest datasets are missing")

    required = {
        "ucsc_enigma_brcamfa_v1_1": multifactorial_source,
        "zanti_2025_case_control": case_control_source,
    }
    for source_id, path in required.items():
        definition = configured.get(source_id)
        if not isinstance(definition, dict):
            raise RuntimeError(f"Clinical LR source manifest lacks {source_id}")
        if not path.is_file():
            raise RuntimeError(f"Required clinical LR source is missing: {path}")
        actual = sha256(path)
        if definition.get("sha256") != actual:
            raise RuntimeError(
                f"Clinical LR source checksum mismatch for {source_id}: "
                f"expected {definition.get('sha256')}, found {actual}"
            )
    return manifest


def parse_values(raw: str) -> list[float | None]:
    if not raw.strip():
        return []
    values = []
    for item in raw.strip().split(","):
        item = item.strip()
        values.append(None if item in {"", "NULL", "NA"} else float(item))
    return values


def strength_for_lr(lr: float) -> tuple[str | None, str | None, int]:
    if lr >= 350:
        return "PP4", "Very Strong", 8
    if lr >= 18.7:
        return "PP4", "Strong", 4
    if lr >= 4.3:
        return "PP4", "Moderate", 2
    if lr >= 2.08:
        return "PP4", "Supporting", 1
    if lr <= 0.00285:
        return "BP5", "Very Strong", -8
    if lr <= 0.05:
        return "BP5", "Strong", -4
    if lr <= 0.23:
        return "BP5", "Moderate", -2
    if lr <= 0.48:
        return "BP5", "Supporting", -1
    return None, None, 0


def apply_combined_evidence(record: dict) -> None:
    """Recompute the combined LR after independently sourced rows are merged."""
    components = sorted(record["source_components"], key=lambda item: item["source_id"])
    source_ids = [component["source_id"] for component in components]
    if len(source_ids) != len(set(source_ids)):
        raise RuntimeError(
            f"Duplicate clinical LR source component for {record['gene']}:"
            f"{record['canonical_c_notation']}"
        )
    independence_groups = [component["independence_group"] for component in components]
    if len(independence_groups) != len(set(independence_groups)):
        raise RuntimeError(
            f"Clinical LR independence group counted more than once for {record['gene']}:"
            f"{record['canonical_c_notation']}"
        )
    combined_lr = math.prod(component["component_lr"] for component in components)
    code, strength, points = strength_for_lr(combined_lr)
    record.update({
        "source_components": components,
        "combined_lr": combined_lr,
        "log10_combined_lr": math.log10(combined_lr) if combined_lr > 0 else None,
        "criterion": code,
        "strength": strength,
        "points": points,
        "informative": code is not None,
    })


def load_indel_reference() -> tuple[dict[str, dict], dict[str, str], dict[str, str]]:
    """Load the indel snapshot as a required, checksum-verified build input."""
    if not INDEL_INDEX.is_file() or not INDEL_METADATA.is_file():
        raise RuntimeError("Normalized BRCA indel snapshot or metadata is missing")
    metadata = json.loads(INDEL_METADATA.read_text(encoding="utf-8"))
    if metadata.get("status") != "validated_reference_snapshot":
        raise RuntimeError("Normalized BRCA indel snapshot is not validated")
    if metadata.get("index_sha256") != sha256(INDEL_INDEX):
        raise RuntimeError("Normalized BRCA indel snapshot checksum mismatch")
    records = json.loads(INDEL_INDEX.read_text(encoding="utf-8"))
    if metadata.get("records") != len(records):
        raise RuntimeError("Normalized BRCA indel snapshot record count mismatch")

    aliases: dict[str, str] = {key: key for key in records}
    ambiguous: set[str] = set()
    for canonical_key, record in records.items():
        for notation in record.get("input_c_notations", []):
            alias = f"{record['gene']}:{notation}"
            previous = aliases.get(alias)
            if previous and previous != canonical_key:
                ambiguous.add(alias)
            else:
                aliases[alias] = canonical_key
    if ambiguous:
        raise RuntimeError(
            f"Normalized BRCA indel snapshot has ambiguous aliases: {len(ambiguous)}"
        )
    return records, aliases, {
        "index_sha256": sha256(INDEL_INDEX),
        "metadata_sha256": sha256(INDEL_METADATA),
        "source_release": str(metadata.get("source_release", "")),
    }


def canonicalize_source_variant(
    gene: str,
    c_notation: str,
    indel_records: dict[str, dict],
    indel_aliases: dict[str, str],
) -> tuple[str, dict[str, str], bool]:
    """Normalize against the local transcript and cross-check known indels.

    The HGVS engine verifies any explicitly supplied deleted or duplicated
    sequence against the checksum-pinned transcript. The indel snapshot is an
    independent cross-check, not a fallback source of an unverified alias.
    """
    from backend.modules.hgvs_engine import derive_protein_consequence

    normalized = derive_protein_consequence(gene, c_notation)
    canonical_c = normalized.canonical_c_notation
    matched_indel = False
    if re.search(r"delins|del|dup|ins", c_notation, re.IGNORECASE):
        submitted_key = f"{gene}:{c_notation}"
        normalized_key = f"{gene}:{canonical_c}"
        indel_key = indel_aliases.get(submitted_key) or indel_aliases.get(normalized_key)
        if indel_key:
            reference_record = indel_records[indel_key]
            reference_c = reference_record["canonical_c_notation"]
            if reference_c != canonical_c:
                raise RuntimeError(
                    "HGVS normalization conflicts with normalized indel snapshot: "
                    f"{gene} {c_notation} -> {canonical_c}, snapshot -> {reference_c}"
                )
            matched_indel = True
    return canonical_c, normalized.provenance, matched_indel


def build(
    source: Path,
    output: Path,
    metadata_path: Path,
    case_control_source: Path = DEFAULT_CASE_CONTROL_SOURCE,
    source_manifest_path: Path = DEFAULT_SOURCE_MANIFEST,
) -> dict:
    source_manifest = load_source_manifest(
        source_manifest_path, source, case_control_source
    )
    indel_records, indel_aliases, indel_dependency = load_indel_reference()
    records: dict[str, dict] = {}
    conflicting_keys: set[str] = set()
    conflicting_records: dict[str, dict] = {}
    excluded: Counter[str] = Counter()
    normalization_counts: Counter[str] = Counter()
    normalization_failures: list[dict[str, str]] = []
    normalization_provenance: dict[str, str] | None = None
    rows_seen = 0
    case_control_rows_seen = 0
    case_control_rows_selected = 0
    case_control_rows_admitted = 0
    case_control_selected_by_gene: Counter[str] = Counter()
    case_control_by_gene: Counter[str] = Counter()
    case_control_unavailable_records: list[dict[str, str]] = []

    with source.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        if reader.fieldnames:
            reader.fieldnames[0] = reader.fieldnames[0].lstrip("#")
        for row in reader:
            rows_seen += 1
            name = row.get("name", "")
            if ":c." not in name:
                excluded["invalid_hgvs"] += 1
                continue
            transcript, c_notation = name.split(":", 1)
            gene = TRANSCRIPTS.get(transcript)
            if not gene:
                excluded["unsupported_transcript"] += 1
                continue

            try:
                canonical_c, provenance, matched_indel = canonicalize_source_variant(
                    gene, c_notation, indel_records, indel_aliases
                )
            except ValueError as exc:
                excluded[f"hgvs_normalization_failed:{type(exc).__name__}"] += 1
                normalization_failures.append({
                    "variant": name,
                    "error_code": str(getattr(exc, "code", type(exc).__name__)),
                    "reason": str(exc),
                })
                continue
            if normalization_provenance is None:
                normalization_provenance = provenance
            elif normalization_provenance != provenance:
                raise RuntimeError("HGVS normalization provenance changed during snapshot build")
            normalization_counts["source_records_normalized"] += 1
            if canonical_c != c_notation:
                normalization_counts["notations_canonicalized"] += 1
            if re.search(r"delins|del|dup|ins", c_notation, re.IGNORECASE):
                normalization_counts[
                    "known_indels_cross_checked" if matched_indel else "indels_not_in_reference_snapshot"
                ] += 1

            components = []
            all_values = []
            for field, definition in SOURCES.items():
                values = parse_values(row.get(field, ""))
                typed_values = [
                    {"data_type": definition["data_types"][idx], "lr": value}
                    for idx, value in enumerate(values[:len(definition["data_types"])])
                    if value is not None
                ]
                if not typed_values:
                    continue
                component_lr = math.prod(value["lr"] for value in typed_values)
                all_values.extend(value["lr"] for value in typed_values)
                components.append({
                    "source_id": definition["source_id"],
                    "citation": definition["citation"],
                    "pmid": definition["pmid"],
                    "doi": definition["doi"],
                    "clinical_data": typed_values,
                    "component_lr": component_lr,
                    "evidence_family": "published_multifactorial_clinical_lr",
                    "independence_group": definition["independence_group"],
                    "source_dataset": "UCSC ENIGMA BRCAmfa track 1.1.0",
                })
            if not all_values:
                excluded["no_appendix_b_lr"] += 1
                continue

            input_notations = {c_notation, canonical_c}
            source_interval = {
                "chrom": row["chrom"].removeprefix("chr"),
                "start_0_based": int(row["chromStart"]),
                "end_0_based": int(row["chromEnd"]),
            }

            key = f"{gene}:{canonical_c}"
            if key in conflicting_keys:
                excluded["conflicting_canonical_record"] += 1
                continue
            record = {
                "gene": gene,
                "reference_transcript": transcript,
                "canonical_c_notation": canonical_c,
                "input_c_notations": sorted(input_notations),
                "source_grch38_intervals": [source_interval],
                "source_components": components,
                "source": {
                    "dataset": "UCSC ENIGMA BRCAmfa track",
                    "track_version": "ENIGMA specifications 1.1.0",
                    "track_url": TRACK_URL,
                    "description_url": TRACK_DESCRIPTION_URL,
                },
            }
            apply_combined_evidence(record)
            previous = records.get(key)
            if previous:
                existing_by_source_id = {
                    component["source_id"]: component
                    for component in previous["source_components"]
                }
                component_conflict = any(
                    component["source_id"] in existing_by_source_id
                    and existing_by_source_id[component["source_id"]] != component
                    for component in record["source_components"]
                )
                if component_conflict:
                    excluded["conflicting_canonical_record"] += 1
                    conflicting_keys.add(key)
                    conflicting_records[key] = {
                        "reason": "different clinical LR components under the same source ID",
                        "existing_input_c_notations": previous["input_c_notations"],
                        "incoming_input_c_notations": record["input_c_notations"],
                        "existing_source_components": previous["source_components"],
                        "incoming_source_components": record["source_components"],
                    }
                    records.pop(key, None)
                    continue
                normalization_counts["canonical_source_rows_merged"] += 1
                previous["input_c_notations"] = sorted(set(
                    previous["input_c_notations"] + record["input_c_notations"]
                ))
                previous["source_grch38_intervals"] = sorted(
                    {
                        (item["chrom"], item["start_0_based"], item["end_0_based"])
                        for item in previous["source_grch38_intervals"]
                        + record["source_grch38_intervals"]
                    }
                )
                previous["source_grch38_intervals"] = [
                    {"chrom": chrom, "start_0_based": start, "end_0_based": end}
                    for chrom, start, end in previous["source_grch38_intervals"]
                ]
                new_components = [
                    component for component in record["source_components"]
                    if component["source_id"] not in existing_by_source_id
                ]
                normalization_counts["independent_source_components_merged"] += len(new_components)
                normalization_counts["duplicate_source_components_deduplicated"] += (
                    len(record["source_components"]) - len(new_components)
                )
                previous["source_components"].extend(new_components)
                apply_combined_evidence(previous)
                continue
            records[key] = record

    workbook = openpyxl.load_workbook(
        case_control_source, read_only=True, data_only=True
    )
    if ZANTI_SHEET not in workbook.sheetnames:
        raise RuntimeError(
            f"Zanti case-control workbook lacks required sheet {ZANTI_SHEET!r}"
        )
    sheet = workbook[ZANTI_SHEET]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=5, max_row=5))]
    expected_columns = {
        0: "hg19",
        1: "hg38",
        9: "GENE",
        12: "HGVScb",
        24: "FAF",
        36: "Ng",
        37: "LR",
        38: "Suggested         ACMG/AMP evidenceh",
        39: "Dataset Directioni",
    }
    for column_index, expected_name in expected_columns.items():
        if header[column_index] != expected_name:
            raise RuntimeError(
                "Unexpected Zanti Supplementary Data 5 schema at column "
                f"{column_index + 1}: expected {expected_name!r}, "
                f"found {header[column_index]!r}"
            )

    for values in sheet.iter_rows(min_row=6, values_only=True):
        gene = str(values[9] or "").strip()
        if gene not in {"BRCA1", "BRCA2"}:
            continue
        case_control_rows_seen += 1
        c_notation = str(values[12] or "").strip()
        suggested_evidence = str(values[38] or "").strip()
        if not c_notation:
            excluded["zanti_missing_hgvsc"] += 1
            continue
        if not suggested_evidence:
            excluded["zanti_missing_recommendation"] += 1
            continue
        if suggested_evidence == "No evidence*":
            excluded["zanti_below_carrier_or_dataset_requirement"] += 1
            continue
        faf = values[24]
        if faf not in {None, ""} and float(faf) > 0.001:
            excluded["zanti_faf_above_0_001"] += 1
            continue
        if suggested_evidence == "N/A":
            excluded["zanti_lr_not_estimable"] += 1
            case_control_unavailable_records.append({
                "gene": gene,
                "c_notation": c_notation,
                "published_lr": str(values[37]),
                "published_evidence_label": suggested_evidence,
                "carriers": str(values[36]),
                "dataset_direction": str(values[39] or "").strip(),
            })
            continue
        case_control_rows_selected += 1
        case_control_selected_by_gene[gene] += 1
        try:
            component_lr = float(values[37])
        except (TypeError, ValueError):
            excluded["zanti_lr_not_estimable"] += 1
            case_control_unavailable_records.append({
                "gene": gene,
                "c_notation": c_notation,
                "published_lr": str(values[37]),
                "published_evidence_label": suggested_evidence,
                "carriers": str(values[36]),
                "dataset_direction": str(values[39] or "").strip(),
            })
            continue
        try:
            canonical_c, provenance, matched_indel = canonicalize_source_variant(
                gene, c_notation, indel_records, indel_aliases
            )
        except ValueError as exc:
            raise RuntimeError(
                "A recommended Zanti 2025 case-control record could not be "
                f"normalized: {gene} {c_notation}: {exc}"
            ) from exc
        if normalization_provenance is None:
            normalization_provenance = provenance
        elif normalization_provenance != provenance:
            raise RuntimeError("HGVS normalization provenance changed during snapshot build")
        normalization_counts["zanti_records_normalized"] += 1
        if canonical_c != c_notation:
            normalization_counts["zanti_notations_canonicalized"] += 1
        if re.search(r"delins|del|dup|ins", c_notation, re.IGNORECASE):
            normalization_counts[
                "zanti_known_indels_cross_checked"
                if matched_indel else "zanti_indels_not_in_reference_snapshot"
            ] += 1

        if not math.isfinite(component_lr) or component_lr < 0:
            raise RuntimeError(
                f"Invalid Zanti case-control LR for {gene} {c_notation}: {values[37]!r}"
            )
        component = {
            "source_id": ZANTI_SOURCE["source_id"],
            "citation": ZANTI_SOURCE["citation"],
            "pmid": ZANTI_SOURCE["pmid"],
            "doi": ZANTI_SOURCE["doi"],
            "clinical_data": [{"data_type": "case_control", "lr": component_lr}],
            "component_lr": component_lr,
            "numeric_note": (
                "published value is 0 after source-data numerical rounding"
                if component_lr == 0 else None
            ),
            "evidence_family": "case_control_likelihood_ratio",
            "independence_group": ZANTI_SOURCE["independence_group"],
            "source_dataset": "Zanti et al. 2025 Supplementary Data 5",
            "cohorts": ZANTI_SOURCE["cohorts"],
            "carriers": str(values[36]),
            "dataset_direction": str(values[39] or "").strip(),
            "published_evidence_label": suggested_evidence,
            "admission_filters": {
                "within_cds_plus_minus_5bp": True,
                "faf_non_founder_lte": 0.001,
                "minimum_combined_carriers": 3,
                "brca2_minimum_datasets": 2,
            },
        }
        source_interval = {
            "hg19_variant_id": str(values[0]),
            "hg38_variant_id": str(values[1]),
        }
        key = f"{gene}:{canonical_c}"
        previous = records.get(key)
        if previous:
            existing = {
                item["source_id"]: item for item in previous["source_components"]
            }
            if component["source_id"] in existing:
                if existing[component["source_id"]] != component:
                    raise RuntimeError(
                        f"Conflicting Zanti case-control rows normalize to {key}"
                    )
                normalization_counts["duplicate_zanti_components_deduplicated"] += 1
            else:
                previous["source_components"].append(component)
            previous["input_c_notations"] = sorted(
                set(previous["input_c_notations"] + [c_notation, canonical_c])
            )
            previous.setdefault("source_variant_ids", []).append(source_interval)
            apply_combined_evidence(previous)
        else:
            record = {
                "gene": gene,
                "reference_transcript": (
                    "NM_007294.4" if gene == "BRCA1" else "NM_000059.4"
                ),
                "canonical_c_notation": canonical_c,
                "input_c_notations": sorted({c_notation, canonical_c}),
                "source_variant_ids": [source_interval],
                "source_components": [component],
                "source": {
                    "dataset": "Zanti et al. 2025 Supplementary Data 5",
                    "source_url": ZANTI_SOURCE_URL,
                },
            }
            apply_combined_evidence(record)
            records[key] = record
        case_control_rows_admitted += 1
        case_control_by_gene[gene] += 1

    if case_control_rows_seen != 6909:
        raise RuntimeError(
            "Unexpected Zanti Supplementary Data 5 BRCA row count: "
            f"expected 6909, found {case_control_rows_seen}"
        )
    if case_control_rows_selected != 1710 or case_control_selected_by_gene != Counter(
        {"BRCA1": 681, "BRCA2": 1029}
    ):
        raise RuntimeError(
            "Zanti case-control admission filters produced an unexpected result: "
            f"{case_control_rows_selected} records, {dict(case_control_selected_by_gene)}"
        )
    if case_control_rows_admitted != 1710 or case_control_by_gene != Counter(
        {"BRCA1": 681, "BRCA2": 1029}
    ):
        raise RuntimeError(
            "Zanti quantitative case-control rows produced an unexpected result: "
            f"{case_control_rows_admitted} records, {dict(case_control_by_gene)}"
        )

    records = dict(sorted(records.items()))
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes((json.dumps(records, indent=2, sort_keys=True) + "\n").encode("utf-8"))
    criteria = Counter(record["criterion"] or "not_informative" for record in records.values())
    evidence_counts = Counter(
        component["source_id"]
        for record in records.values()
        for component in record["source_components"]
    )
    metadata = {
        "dataset": "BRCA1/2 combined clinical likelihood-ratio snapshot",
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "status": "validated_derived_snapshot",
        "source_manifest_file": source_manifest_path.name,
        "source_manifest_sha256": sha256(source_manifest_path),
        "source_manifest": source_manifest,
        "source_files": {
            "ucsc_enigma_brcamfa_v1_1": {
                "file": source.name,
                "sha256": sha256(source),
                "url": TRACK_URL,
                "description_url": TRACK_DESCRIPTION_URL,
                "track_version": "ENIGMA specifications 1.1.0",
            },
            "zanti_2025_case_control": {
                "file": case_control_source.name,
                "sha256": sha256(case_control_source),
                "url": ZANTI_SOURCE_URL,
                "sheet": ZANTI_SHEET,
                "pmid": ZANTI_SOURCE["pmid"],
                "doi": ZANTI_SOURCE["doi"],
            },
        },
        "target_rule_version": "ENIGMA BRCA1/2 VCEP 1.2 PP4/BP5 thresholds",
        "reference_transcripts": TRANSCRIPTS,
        "rows_seen": rows_seen,
        "records": len(records),
        "criteria": dict(sorted(criteria.items())),
        "records_by_source_id": dict(sorted(evidence_counts.items())),
        "included_multifactorial_sources": SOURCES,
        "included_case_control_source": ZANTI_SOURCE,
        "case_control_source_rows": case_control_rows_seen,
        "case_control_records_selected": case_control_rows_selected,
        "case_control_selected_by_gene": dict(sorted(case_control_selected_by_gene.items())),
        "case_control_quantitative_records": case_control_rows_admitted,
        "case_control_quantitative_records_by_gene": dict(sorted(case_control_by_gene.items())),
        "case_control_unavailable_records": case_control_unavailable_records,
        "combination_policy": {
            "method": "multiply admitted variant-specific clinical LR components",
            "single_final_code": True,
            "component_codes_scored_separately": False,
            "duplicate_source_ids": "fatal",
            "duplicate_independence_groups": "fatal",
            "zanti_admission": (
                "Supplementary Data 5; CDS +/-5 bp; non-founder FAF <=0.001; "
                ">=3 combined carriers; BRCA2 evidence from >=2 datasets"
            ),
        },
        "normalization": {
            "method": "biocommons.hgvs with checksum-pinned cdot panel provider",
            "provenance": normalization_provenance or {},
            "counts": dict(sorted(normalization_counts.items())),
            "normalized_indel_dependency": indel_dependency,
            "failures": normalization_failures,
        },
        "excluded": dict(sorted(excluded.items())),
        "conflicting_canonical_keys": sorted(conflicting_keys),
        "conflicting_canonical_records": dict(sorted(conflicting_records.items())),
        "index_sha256": sha256(output),
    }
    metadata_path.write_bytes((json.dumps(metadata, indent=2, sort_keys=True) + "\n").encode("utf-8"))
    return metadata


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    parser.add_argument(
        "--case-control-source", type=Path, default=DEFAULT_CASE_CONTROL_SOURCE
    )
    parser.add_argument(
        "--source-manifest", type=Path, default=DEFAULT_SOURCE_MANIFEST
    )
    args = parser.parse_args()
    print(json.dumps(build(
        args.source,
        args.output,
        args.metadata,
        args.case_control_source,
        args.source_manifest,
    ), indent=2))


if __name__ == "__main__":
    main()
